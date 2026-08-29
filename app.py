import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from io import BytesIO
import re
import difflib

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# PAGE
# =========================================================

st.set_page_config(
    page_title="GST Reconciliation Pro",
    page_icon="🧾",
    layout="wide"
)


# =========================================================
# CSS
# =========================================================

st.markdown("""
<style>
.stApp {
    background-color: #f4f7fb;
}

[data-testid="stHeader"] {
    background-color: transparent;
}

.hero {
    background: linear-gradient(135deg, #0f2a5f, #1769aa);
    padding: 30px;
    border-radius: 20px;
    margin-bottom: 25px;
}

.hero h1 {
    color: white !important;
    font-size: 34px;
    margin: 0;
}

.hero p {
    color: #dbeafe !important;
    margin-top: 8px;
}

.section-title {
    color: #0f2a5f;
    font-size: 23px;
    font-weight: 800;
    margin-top: 25px;
    margin-bottom: 15px;
}

.kpi {
    background: white;
    border-radius: 16px;
    padding: 20px;
    border: 1px solid #dbe3ef;
    box-shadow: 0 5px 18px rgba(15,42,95,.07);
    min-height: 125px;
}

.kpi-title {
    color: #64748b;
    font-size: 12px;
    font-weight: 800;
    letter-spacing: .03em;
}

.kpi-value {
    color: #172033;
    font-size: 30px;
    font-weight: 900;
    margin-top: 8px;
}

.kpi-desc {
    color: #94a3b8;
    font-size: 12px;
}

.blue {
    border-top: 4px solid #2563eb;
}

.green {
    border-top: 4px solid #16a34a;
}

.red {
    border-top: 4px solid #dc2626;
}

.orange {
    border-top: 4px solid #ea580c;
}

.purple {
    border-top: 4px solid #7c3aed;
}

.teal {
    border-top: 4px solid #0d9488;
}

.info-box {
    background: white;
    padding: 18px;
    border-radius: 13px;
    border-left: 5px solid #2563eb;
    border-top: 1px solid #dbe3ef;
    border-right: 1px solid #dbe3ef;
    border-bottom: 1px solid #dbe3ef;
    margin: 15px 0;
}

.warn-box {
    background: #fff7ed;
    padding: 14px 18px;
    border-radius: 13px;
    border-left: 5px solid #ea580c;
    margin: 10px 0;
    color: #7c2d12;
    font-size: 13px;
}

[data-testid="stFileUploaderDropzone"] {
    background: white;
    border: 2px dashed #b8c9df;
    border-radius: 14px;
}

.stButton button {
    border-radius: 10px;
    font-weight: 700;
}

.stDownloadButton button {
    background: #0f2a5f !important;
    color: white !important;
    border-radius: 10px;
    font-weight: 700;
}

.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 999px;
    font-size: 11px;
    font-weight: 700;
}

/* =========================================================
   YOOM-STYLE ANIMATIONS
   ========================================================= */

@keyframes floatUpDown {
    0%, 100% { transform: translateY(0px); }
    50% { transform: translateY(-14px); }
}
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(18px); }
    to { opacity: 1; transform: translateY(0); }
}
@keyframes gradientShift {
    0% { background-position: 0% 50%; }
    50% { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}
@keyframes blinkCursor {
    0%, 50% { opacity: 1; }
    51%, 100% { opacity: 0; }
}

.hero {
    background: linear-gradient(-45deg, #0f2a5f, #1769aa, #2563eb, #0f2a5f);
    background-size: 300% 300%;
    animation: gradientShift 8s ease infinite;
    position: relative;
    overflow: hidden;
}
.hero::before, .hero::after {
    content: "";
    position: absolute;
    border-radius: 50%;
    background: rgba(255,255,255,0.08);
}
.hero::before {
    width: 180px; height: 180px;
    top: -60px; right: 40px;
    animation: floatUpDown 6s ease-in-out infinite;
}
.hero::after {
    width: 110px; height: 110px;
    bottom: -30px; left: 60px;
    animation: floatUpDown 5s ease-in-out infinite reverse;
}
.hero h1 { animation: fadeInUp 0.8s ease; }
.hero p { animation: fadeInUp 0.8s ease 0.2s backwards; }
.hero-cursor {
    display: inline-block;
    width: 3px; height: 1em;
    background: #fff;
    margin-left: 5px;
    animation: blinkCursor 1s step-start infinite;
    vertical-align: middle;
}

.float-icons {
    display: flex;
    gap: 18px;
    justify-content: center;
    margin: 22px 0 30px;
    flex-wrap: wrap;
}
.float-icon {
    width: 56px; height: 56px;
    border-radius: 50%;
    background: white;
    display: flex; align-items: center; justify-content: center;
    font-size: 24px;
    box-shadow: 0 6px 18px rgba(15,42,95,.12);
    animation: floatUpDown 3.5s ease-in-out infinite;
}
.float-icon:nth-child(2) { animation-delay: .3s; }
.float-icon:nth-child(3) { animation-delay: .6s; }
.float-icon:nth-child(4) { animation-delay: .9s; }
.float-icon:nth-child(5) { animation-delay: 1.2s; }
.float-icon:nth-child(6) { animation-delay: 1.5s; }

.feature-card {
    background: white;
    border-radius: 16px;
    padding: 18px 20px;
    border: 1px solid #dbe3ef;
    box-shadow: 0 5px 18px rgba(15,42,95,.06);
    opacity: 0;
    animation: fadeInUp 0.6s ease forwards;
    min-height: 150px;
    transition: transform .2s ease;
}
.feature-card:hover { transform: translateY(-4px); }
.feature-card:nth-child(1) { animation-delay: .1s; }
.feature-card:nth-child(2) { animation-delay: .25s; }
.feature-card:nth-child(3) { animation-delay: .4s; }
.feature-card:nth-child(4) { animation-delay: .55s; }

/* =========================================================
   LOADING MASCOTS (Yoom-style cute characters)
   ========================================================= */

@keyframes mascotBounce {
    0%, 100% { transform: translateY(0) rotate(0deg); }
    50% { transform: translateY(-22px) rotate(-6deg); }
}
@keyframes mascotBounce2 {
    0%, 100% { transform: translateY(0) rotate(0deg); }
    50% { transform: translateY(-16px) rotate(6deg); }
}
@keyframes mascotWave {
    0%, 100% { transform: rotate(0deg); }
    25% { transform: rotate(-10deg); }
    75% { transform: rotate(10deg); }
}
@keyframes textPulse {
    0%, 100% { opacity: 0.6; }
    50% { opacity: 1; }
}

.loading-mascot {
    background: linear-gradient(135deg, #eef4ff, #f4f7fb);
    border-radius: 20px;
    padding: 34px 20px;
    text-align: center;
    border: 1px solid #dbe3ef;
    margin: 10px 0 18px;
}
.mascot-row {
    display: flex;
    justify-content: center;
    align-items: flex-end;
    gap: 26px;
    margin-bottom: 14px;
}
.mascot-char {
    width: 64px; height: 64px;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 30px;
    box-shadow: 0 8px 20px rgba(15,42,95,.15);
}
.mascot-1 { background: #fde68a; animation: mascotBounce 1.6s ease-in-out infinite; }
.mascot-2 { background: #fbcfe8; animation: mascotWave 1.4s ease-in-out infinite; animation-delay: .15s; }
.mascot-3 { background: #a7f3d0; animation: mascotBounce2 1.8s ease-in-out infinite; animation-delay: .3s; }
.mascot-4 { background: #bfdbfe; animation: mascotBounce 2s ease-in-out infinite; animation-delay: .45s; }
.loading-text {
    color: #0f2a5f;
    font-weight: 700;
    font-size: 15px;
    animation: textPulse 1.4s ease-in-out infinite;
    margin: 0;
}

/* =========================================================
   HERO SIDE MASCOT (waving character next to title)
   ========================================================= */

.hero-flex {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 20px;
}
.hero-text-block {
    flex: 1;
}
.hero-mascot {
    width: 92px; height: 92px;
    border-radius: 50%;
    background: rgba(255,255,255,0.15);
    display: flex; align-items: center; justify-content: center;
    font-size: 46px;
    flex-shrink: 0;
    animation: floatUpDown 3s ease-in-out infinite;
    position: relative;
}
.hero-mascot .hero-hand {
    position: absolute;
    top: -6px; right: -6px;
    font-size: 26px;
    animation: mascotWave 1.1s ease-in-out infinite;
    transform-origin: 70% 70%;
}

/* =========================================================
   RUNNING LOADER OVERLAY (blur background + running character)
   ========================================================= */

@keyframes runCycle {
    0%   { transform: translateX(-120px) scaleX(1); }
    48%  { transform: translateX(90px) scaleX(1); }
    50%  { transform: translateX(90px) scaleX(-1); }
    98%  { transform: translateX(-120px) scaleX(-1); }
    100% { transform: translateX(-120px) scaleX(1); }
}
@keyframes legBounce {
    0%, 100% { transform: translateY(0); }
    50% { transform: translateY(-6px); }
}
@keyframes blurIn {
    from { backdrop-filter: blur(0px); opacity: 0; }
    to { backdrop-filter: blur(6px); opacity: 1; }
}

.run-overlay {
    background: rgba(244,247,251,0.75);
    animation: blurIn 0.5s ease forwards;
    border-radius: 20px;
    padding: 46px 20px;
    text-align: center;
    border: 1px solid #dbe3ef;
    margin: 10px 0 18px;
    overflow: hidden;
    position: relative;
}
.run-track {
    position: relative;
    height: 70px;
    display: flex;
    align-items: center;
    justify-content: center;
}
.runner {
    font-size: 44px;
    display: inline-block;
    animation: runCycle 2.2s linear infinite, legBounce 0.3s ease-in-out infinite;
}
.run-ground {
    width: 260px;
    height: 3px;
    background: repeating-linear-gradient(90deg, #b8c9df 0 14px, transparent 14px 28px);
    margin: 6px auto 0;
    opacity: 0.6;
}

/* =========================================================
   AI WORKER UI — playful, premium & professional
   ========================================================= */

@keyframes workerFloat {
    0%, 100% { transform: translateY(0) rotate(0deg); }
    50% { transform: translateY(-10px) rotate(-1.5deg); }
}
@keyframes workerFloatAlt {
    0%, 100% { transform: translateY(0) rotate(0deg); }
    50% { transform: translateY(-14px) rotate(1.5deg); }
}
@keyframes handWave2 {
    0%, 100% { transform: rotate(8deg); }
    50% { transform: rotate(-24deg); }
}
@keyframes eyeBlink {
    0%, 92%, 100% { transform: scaleY(1); }
    95% { transform: scaleY(.12); }
}
@keyframes chipDrift {
    0%, 100% { transform: translateY(0) translateX(0); }
    50% { transform: translateY(-7px) translateX(4px); }
}
@keyframes stageGlow {
    0%, 100% { opacity: .45; transform: scale(.96); }
    50% { opacity: .85; transform: scale(1.02); }
}
@keyframes processingDots {
    0%, 20% { opacity: .25; transform: translateY(0); }
    50% { opacity: 1; transform: translateY(-3px); }
    80%, 100% { opacity: .25; transform: translateY(0); }
}
@keyframes successPop {
    0% { transform: scale(.7); opacity: 0; }
    70% { transform: scale(1.08); opacity: 1; }
    100% { transform: scale(1); opacity: 1; }
}

.hero {
    min-height: 300px;
    border-radius: 24px;
    padding: 28px 32px 18px !important;
    box-shadow: 0 18px 55px rgba(15,42,95,.16);
}
.hero::before {
    width: 260px !important;
    height: 260px !important;
    top: -115px !important;
    right: 8% !important;
}
.hero::after {
    width: 180px !important;
    height: 180px !important;
    bottom: -100px !important;
    left: 8% !important;
}
.hero-topline {
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:18px;
    position:relative;
    z-index:3;
}
.hero-kicker {
    display:inline-flex;
    align-items:center;
    gap:7px;
    padding:6px 11px;
    border:1px solid rgba(255,255,255,.24);
    background:rgba(255,255,255,.10);
    border-radius:999px;
    color:#eaf3ff;
    font-size:11px;
    font-weight:800;
    letter-spacing:.06em;
    text-transform:uppercase;
}
.hero-live {
    display:inline-flex;
    align-items:center;
    gap:7px;
    color:#dfffea;
    font-size:11px;
    font-weight:800;
}
.hero-live-dot {
    width:8px;
    height:8px;
    border-radius:50%;
    background:#4ade80;
    box-shadow:0 0 0 5px rgba(74,222,128,.13);
    animation:stageGlow 1.8s ease-in-out infinite;
}
.hero-main-title {
    position:relative;
    z-index:3;
    margin-top:13px;
}
.hero-main-title h1 {
    font-size:42px !important;
    line-height:1.04 !important;
    letter-spacing:-.035em;
    margin:0 !important;
}
.hero-main-title p {
    max-width:760px;
    font-size:14px !important;
}
.hero-pills {
    display:flex;
    gap:8px;
    flex-wrap:wrap;
    margin-top:15px;
    position:relative;
    z-index:3;
}
.hero-pill {
    padding:6px 10px;
    border-radius:999px;
    background:rgba(255,255,255,.12);
    border:1px solid rgba(255,255,255,.18);
    color:#f8fbff;
    font-size:11px;
    font-weight:700;
    backdrop-filter:blur(8px);
}
.ai-stage {
    position:relative;
    height:118px;
    margin-top:10px;
    overflow:hidden;
    z-index:2;
}
.ai-stage-glow {
    position:absolute;
    left:12%;
    right:12%;
    bottom:7px;
    height:44px;
    border-radius:50%;
    background:rgba(255,255,255,.16);
    filter:blur(18px);
    animation:stageGlow 4s ease-in-out infinite;
}
.worker {
    position:absolute;
    bottom:7px;
    width:74px;
    height:82px;
    border-radius:38% 38% 30% 30%;
    box-shadow:0 13px 24px rgba(7,22,49,.18), inset 0 -8px 0 rgba(0,0,0,.06);
    animation:workerFloat 3.4s ease-in-out infinite;
}
.worker::before {
    content:"";
    position:absolute;
    width:42px;
    height:18px;
    left:16px;
    bottom:-8px;
    border-radius:50%;
    background:rgba(7,22,49,.18);
    filter:blur(5px);
    z-index:-1;
}
.worker .worker-eyes {
    position:absolute;
    top:24px;
    left:20px;
    right:20px;
    display:flex;
    justify-content:space-between;
}
.worker .worker-eyes span {
    width:7px;
    height:10px;
    border-radius:50%;
    background:#172033;
    animation:eyeBlink 4.5s ease-in-out infinite;
}
.worker .worker-mouth {
    position:absolute;
    left:27px;
    top:43px;
    width:20px;
    height:10px;
    border-bottom:3px solid #172033;
    border-radius:0 0 18px 18px;
}
.worker .worker-arm {
    position:absolute;
    width:13px;
    height:39px;
    border-radius:12px;
    background:inherit;
    top:34px;
}
.worker .left-arm { left:-9px; transform:rotate(18deg); }
.worker .right-arm { right:-9px; transform:rotate(-18deg); }
.worker .worker-hand {
    position:absolute;
    right:-21px;
    top:9px;
    width:18px;
    height:18px;
    border-radius:50%;
    background:inherit;
    animation:handWave2 1.15s ease-in-out infinite;
    transform-origin:55% 85%;
}
.worker-yellow { left:27%; background:linear-gradient(145deg,#fde68a,#fbbf24); animation-delay:.05s; }
.worker-pink { left:37%; background:linear-gradient(145deg,#fbcfe8,#f472b6); animation:workerFloatAlt 3.7s ease-in-out infinite; animation-delay:.22s; }
.worker-blue { left:48%; background:linear-gradient(145deg,#93c5fd,#2563eb); animation-delay:.42s; }
.worker-green { left:59%; background:linear-gradient(145deg,#86efac,#10b981); animation:workerFloatAlt 3.5s ease-in-out infinite; animation-delay:.64s; }

.worker-badge {
    position:absolute;
    top:-18px;
    left:50%;
    transform:translateX(-50%);
    white-space:nowrap;
    padding:4px 8px;
    border-radius:999px;
    background:rgba(255,255,255,.93);
    color:#172033;
    font-size:9px;
    font-weight:900;
    box-shadow:0 5px 16px rgba(7,22,49,.12);
}
.ai-chip {
    position:absolute;
    padding:6px 9px;
    border-radius:10px;
    background:rgba(255,255,255,.9);
    color:#172033;
    font-size:9px;
    font-weight:900;
    box-shadow:0 6px 18px rgba(7,22,49,.12);
    animation:chipDrift 3.2s ease-in-out infinite;
    backdrop-filter:blur(7px);
}
.ai-chip.one { left:13%; top:24px; animation-delay:.2s; }
.ai-chip.two { left:72%; top:13px; animation-delay:.9s; }
.ai-chip.three { left:80%; top:67px; animation-delay:1.5s; }
.ai-chip.four { left:18%; top:70px; animation-delay:1.1s; }

.kpi {
    transition:transform .22s ease, box-shadow .22s ease, border-color .22s ease;
}
.kpi:hover {
    transform:translateY(-5px);
    box-shadow:0 13px 30px rgba(15,42,95,.12);
    border-color:#c6d5e8;
}
.stButton button, .stDownloadButton button {
    transition:transform .18s ease, box-shadow .18s ease, filter .18s ease;
}
.stButton button:hover, .stDownloadButton button:hover {
    transform:translateY(-2px);
    box-shadow:0 8px 20px rgba(15,42,95,.14);
    filter:brightness(1.03);
}
[data-testid="stFileUploaderDropzone"] {
    transition:transform .2s ease, border-color .2s ease, box-shadow .2s ease;
}
[data-testid="stFileUploaderDropzone"]:hover {
    transform:translateY(-2px);
    border-color:#6d9de8;
    box-shadow:0 10px 26px rgba(37,99,235,.10);
}
.section-title {
    position:relative;
}
.section-title::after {
    content:"";
    display:block;
    width:52px;
    height:3px;
    margin-top:7px;
    border-radius:99px;
    background:linear-gradient(90deg,#2563eb,#7c3aed);
}
.worker-loader {
    display:flex;
    align-items:center;
    justify-content:center;
    gap:28px;
    height:82px;
}
.loader-worker {
    position:relative;
    width:54px;
    height:62px;
    border-radius:38% 38% 30% 30%;
    box-shadow:0 9px 18px rgba(15,42,95,.16);
    animation:workerFloat 1.1s ease-in-out infinite;
}
.loader-worker:nth-child(2) { animation-delay:.16s; }
.loader-worker:nth-child(3) { animation-delay:.32s; }
.loader-worker::before {
    content:"";
    position:absolute;
    top:19px;
    left:13px;
    width:7px;
    height:9px;
    border-radius:50%;
    background:#172033;
    box-shadow:21px 0 0 #172033;
    animation:eyeBlink 2.4s ease-in-out infinite;
}
.loader-worker::after {
    content:"";
    position:absolute;
    left:20px;
    top:38px;
    width:15px;
    height:7px;
    border-bottom:2px solid #172033;
    border-radius:0 0 12px 12px;
}
.loader-yellow { background:#fbbf24; }
.loader-blue { background:#3b82f6; }
.loader-green { background:#10b981; }
.loading-dots span {
    display:inline-block;
    margin:0 2px;
    animation:processingDots 1.2s ease-in-out infinite;
}
.loading-dots span:nth-child(2){animation-delay:.18s}
.loading-dots span:nth-child(3){animation-delay:.36s}

@media (max-width: 900px) {
    .hero { padding:24px 20px 14px !important; }
    .hero-main-title h1 { font-size:32px !important; }
    .ai-stage { height:108px; }
    .worker-yellow { left:22%; }
    .worker-pink { left:34%; }
    .worker-blue { left:47%; }
    .worker-green { left:60%; }
}
@media (max-width: 600px) {
    .hero-main-title h1 { font-size:27px !important; }
    .hero-topline { align-items:flex-start; }
    .hero-live { display:none; }
    .ai-stage { height:100px; }
    .worker { width:56px; height:65px; }
    .worker .worker-eyes { top:19px; left:15px; right:15px; }
    .worker .worker-mouth { top:35px; left:20px; width:16px; }
    .worker .worker-arm { height:31px; top:29px; }
    .worker-yellow { left:14%; }
    .worker-pink { left:31%; }
    .worker-blue { left:48%; }
    .worker-green { left:65%; }
    .ai-chip { display:none; }
}
@media (prefers-reduced-motion: reduce) {
    *, *::before, *::after {
        animation-duration:.01ms !important;
        animation-iteration-count:1 !important;
        scroll-behavior:auto !important;
    }
}

</style>
""", unsafe_allow_html=True)


# =========================================================
# CONSTANTS
# =========================================================

REQUIRED_FIELDS = ["GSTIN", "Invoice", "Party", "Taxable"]

OPTIONAL_FIELDS = [
    "Date",
    "IGST",
    "CGST",
    "SGST",
    "InvoiceValue"
]

STATUS_COLORS = {
    "Matched": "16A34A",
    "Missing in 2B": "DC2626",
    "Missing in Books": "EA580C",
    "Value Mismatch": "7C3AED",
}


# =========================================================
# HELPERS
# =========================================================

def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def clean_gstin(value):
    if pd.isna(value):
        return ""
    return str(value).upper().replace(" ", "").strip()


def clean_invoice(value):
    if pd.isna(value):
        return ""

    value = str(value).upper().strip()

    if value.endswith(".0"):
        value = value[:-2]

    return re.sub(r"[^A-Z0-9]", "", value)


def clean_amount(value):
    if pd.isna(value):
        return 0.0

    try:
        text = str(value)
        text = text.replace(",", "")
        text = text.replace("₹", "")
        text = text.replace("Rs.", "")
        text = text.replace("Rs", "")
        text = text.strip()

        if text in ("", "-", "nan"):
            return 0.0

        return float(text)

    except Exception:
        return 0.0


def is_valid_gstin(gstin):
    if not gstin or len(gstin) != 15:
        return False

    pattern = (
        r"^[0-9]{2}[A-Z]{5}[0-9]{4}"
        r"[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$"
    )

    return bool(re.match(pattern, gstin))


# =========================================================
# COLUMN ALIASES
# =========================================================

ALIASES = {
    "GSTIN": [
        "GSTIN",
        "GSTIN/UIN",
        "GSTIN of Supplier",
        "GSTIN of supplier",
        "Supplier GSTIN",
        "GST Number",
        "GST No",
        "GST No.",
        "GSTIN of Vendor",
        "GSTIN/UIN of Supplier",
        "Vendor GSTIN"
    ],

    "Invoice": [
        "Invoice Number",
        "Invoice number",
        "Invoice No",
        "Invoice No.",
        "Document Number",
        "Document No",
        "Bill No",
        "Bill Number",
        "Invoice"
    ],

    "Date": [
        "Invoice Date",
        "Invoice date",
        "Document Date",
        "Document date",
        "Bill Date",
        "Date"
    ],

    "Party": [
        "Party Name",
        "Supplier Name",
        "Supplier",
        "Vendor Name",
        "Vendor",
        "Trade Name",
        "Legal Name",
        "Party",
        "Supplier Trade Name",
        "Trade/Legal Name"
    ],

    "Taxable": [
        "Taxable Value",
        "Taxable value",
        "Taxable Amount",
        "Taxable Amount (₹)",
        "Taxable Amt",
        "Taxable"
    ],

    "IGST": [
        "IGST",
        "IGST Amount",
        "IGST Amt",
        "Integrated Tax",
        "Integrated Tax Amount",
        "Integrated Tax Amount (₹)"
    ],

    "CGST": [
        "CGST",
        "CGST Amount",
        "CGST Amt",
        "Central Tax",
        "Central Tax Amount",
        "Central Tax Amount (₹)"
    ],

    "SGST": [
        "SGST",
        "SGST Amount",
        "SGST Amt",
        "UTGST",
        "State Tax",
        "State Tax Amount",
        "State/UT Tax",
        "State/UT Tax Amount",
        "State/UT Tax Amount (₹)"
    ],

    "InvoiceValue": [
        "Invoice Value",
        "Invoice value",
        "Total Invoice Value",
        "Total Value",
        "Document Value",
        "Invoice Amount",
        "Total Invoice Amount",
        "Invoice Value (₹)"
    ]
}


FIELD_LABELS = {
    "GSTIN": "GSTIN",
    "Invoice": "Invoice Number",
    "Date": "Invoice Date",
    "Party": "Party / Supplier Name",
    "Taxable": "Taxable Value",
    "IGST": "IGST",
    "CGST": "CGST",
    "SGST": "SGST",
    "InvoiceValue": "Invoice Value (Total)",
}


def normalize_column(value):
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def find_column(df, aliases):
    columns = {
        normalize_column(c): c
        for c in df.columns
    }

    for alias in aliases:
        key = normalize_column(alias)

        if key in columns:
            return columns[key]

    for column in df.columns:
        current = normalize_column(column)

        for alias in aliases:
            target = normalize_column(alias)

            if target in current or current in target:
                return column

    return None


def auto_detect(df):
    return {
        key: find_column(df, ALIASES[key])
        for key in ALIASES
    }


# =========================================================
# HEADER ROW AUTO-DETECTION
# =========================================================

def detect_header_row(raw_df, max_scan_rows=15):
    """
    Existing header detection logic preserved.
    Scans first rows and finds a row containing
    both GSTIN and Invoice style headers.
    """

    scan_limit = min(max_scan_rows, len(raw_df))

    for row_idx in range(scan_limit):

        row_values = raw_df.iloc[row_idx].tolist()

        norm_values = [
            normalize_column(v)
            for v in row_values
            if pd.notna(v)
        ]

        if not norm_values:
            continue

        has_gstin = any(
            any(
                normalize_column(alias) in val
                or val in normalize_column(alias)
                for alias in ALIASES["GSTIN"]
            )
            for val in norm_values
        )

        has_invoice = any(
            any(
                normalize_column(alias) in val
                or val in normalize_column(alias)
                for alias in ALIASES["Invoice"]
            )
            for val in norm_values
        )

        if has_gstin and has_invoice:
            return row_idx

    return None


# =========================================================
# EXISTING EXCEL READER
# =========================================================

def read_excel(file):
    excel = pd.ExcelFile(file)

    all_data = []
    valid_sheets = []
    skipped_sheets = []

    for sheet in excel.sheet_names:

        try:

            raw = pd.read_excel(
                file,
                sheet_name=sheet,
                header=None
            )

            if raw.empty:
                skipped_sheets.append(sheet)
                continue

            header_row_idx = detect_header_row(raw)

            if header_row_idx is None:
                skipped_sheets.append(sheet)
                continue

            df = pd.read_excel(
                file,
                sheet_name=sheet,
                header=header_row_idx
            )

            df = df.dropna(
                axis=1,
                how="all"
            )

            df = df.loc[
                :,
                ~df.columns.astype(str).str.startswith("Unnamed")
            ]

            gstin_column = find_column(
                df,
                ALIASES["GSTIN"]
            )

            invoice_column = find_column(
                df,
                ALIASES["Invoice"]
            )

            if gstin_column and invoice_column:

                df["_SOURCE_SHEET"] = sheet

                all_data.append(df)
                valid_sheets.append(sheet)

            else:
                skipped_sheets.append(sheet)

        except Exception:
            skipped_sheets.append(sheet)
            continue

    if not all_data:
        raise ValueError(
            "GSTIN aur Invoice Number wale columns kisi bhi sheet mein nahi mile."
        )

    combined = pd.concat(
        all_data,
        ignore_index=True
    )

    return combined, valid_sheets, skipped_sheets


# =========================================================
# NEW PORTAL-SPECIFIC READER
# EXISTING read_excel() IS NOT TOUCHED
# =========================================================

def portal_sheet_type(sheet_name):
    """
    Detects GST portal 2B sheet type from sheet name.
    """

    name = normalize_column(sheet_name)

    if "B2B" in name:
        return "B2B"

    if (
        "CDN" in name
        or "CREDITDEBIT" in name
        or "CREDITNOTE" in name
        or "DEBITNOTE" in name
    ):
        return "CDN"

    if (
        "RCM" in name
        or "REVERSECHARGE" in name
    ):
        return "RCM"

    if "ITCSUMMARY" in name:
        return "ITC Summary"

    if "SUMMARY" in name:
        return "Summary"

    if "HSN" in name:
        return "HSN"

    return "Other"


def read_portal_2b(file):
    """
    Separate portal reader.

    IMPORTANT:
    Existing read_excel() remains untouched.

    Portal logic:
    - B2B sheet is the primary 2B reconciliation source.
    - CDN / credit-debit note sheets are stored separately.
    - RCM sheets are stored separately.
    - ITC Summary and other portal sheets are not mixed into B2B.
    """

    excel = pd.ExcelFile(file)

    b2b_frames = []
    cdn_frames = []
    rcm_frames = []

    valid_b2b_sheets = []
    cdn_sheets = []
    rcm_sheets = []
    skipped_sheets = []

    for sheet in excel.sheet_names:

        try:

            sheet_type = portal_sheet_type(sheet)

            raw = pd.read_excel(
                file,
                sheet_name=sheet,
                header=None
            )

            if raw.empty:
                skipped_sheets.append(sheet)
                continue

            header_row_idx = detect_header_row(raw)

            if header_row_idx is None:

                # Some portal sheets may have slightly different
                # structures. Try a normal header read as fallback.
                try:
                    temp = pd.read_excel(
                        file,
                        sheet_name=sheet
                    )
                except Exception:
                    temp = pd.DataFrame()

                if temp.empty:
                    skipped_sheets.append(sheet)
                    continue

                df = temp

            else:

                df = pd.read_excel(
                    file,
                    sheet_name=sheet,
                    header=header_row_idx
                )

            df = df.dropna(
                axis=1,
                how="all"
            )

            df = df.loc[
                :,
                ~df.columns.astype(str).str.startswith("Unnamed")
            ]

            df["_SOURCE_SHEET"] = sheet

            gstin_column = find_column(
                df,
                ALIASES["GSTIN"]
            )

            invoice_column = find_column(
                df,
                ALIASES["Invoice"]
            )

            # -------------------------------------------------
            # B2B
            # -------------------------------------------------

            if sheet_type == "B2B":

                if gstin_column and invoice_column:
                    b2b_frames.append(df)
                    valid_b2b_sheets.append(sheet)
                else:
                    skipped_sheets.append(sheet)

            # -------------------------------------------------
            # CDN
            # -------------------------------------------------

            elif sheet_type == "CDN":

                cdn_frames.append(df)
                cdn_sheets.append(sheet)

            # -------------------------------------------------
            # RCM
            # -------------------------------------------------

            elif sheet_type == "RCM":

                rcm_frames.append(df)
                rcm_sheets.append(sheet)

            else:

                skipped_sheets.append(sheet)

        except Exception:
            skipped_sheets.append(sheet)

    # ---------------------------------------------------------
    # If no explicit B2B sheet found, fallback to existing reader
    # ---------------------------------------------------------

    if not b2b_frames:

        fallback_raw, fallback_sheets, fallback_skipped = read_excel(file)

        b2b_raw = fallback_raw

        valid_b2b_sheets = fallback_sheets

        skipped_sheets.extend(fallback_skipped)

    else:

        b2b_raw = pd.concat(
            b2b_frames,
            ignore_index=True
        )

    # ---------------------------------------------------------
    # CDN
    # ---------------------------------------------------------

    if cdn_frames:

        cdn_raw = pd.concat(
            cdn_frames,
            ignore_index=True
        )

    else:

        cdn_raw = pd.DataFrame()

    # ---------------------------------------------------------
    # RCM
    # ---------------------------------------------------------

    if rcm_frames:

        rcm_raw = pd.concat(
            rcm_frames,
            ignore_index=True
        )

    else:

        rcm_raw = pd.DataFrame()

    return {
        "b2b_raw": b2b_raw,
        "cdn_raw": cdn_raw,
        "rcm_raw": rcm_raw,
        "b2b_sheets": valid_b2b_sheets,
        "cdn_sheets": cdn_sheets,
        "rcm_sheets": rcm_sheets,
        "skipped_sheets": skipped_sheets,
        "all_sheets": excel.sheet_names
    }


# =========================================================
# STANDARDIZE
# =========================================================

def standardize(df, overrides=None):

    overrides = overrides or {}

    result = pd.DataFrame()
    detected = {}

    for key in ALIASES:

        manual = overrides.get(key)

        if (
            manual
            and manual != "— None —"
            and manual in df.columns
        ):
            detected[key] = manual

        else:
            detected[key] = find_column(
                df,
                ALIASES[key]
            )

    # GSTIN

    if detected["GSTIN"]:
        result["GSTIN"] = df[
            detected["GSTIN"]
        ].apply(clean_gstin)
    else:
        result["GSTIN"] = ""

    # Invoice

    if detected["Invoice"]:
        result["Invoice Number"] = df[
            detected["Invoice"]
        ].apply(clean_invoice)
    else:
        result["Invoice Number"] = ""

    # Date

    if detected["Date"]:
        result["Invoice Date"] = pd.to_datetime(
            df[detected["Date"]],
            errors="coerce"
        )
    else:
        result["Invoice Date"] = pd.NaT

    # Party

    if detected["Party"]:
        result["Party Name"] = df[
            detected["Party"]
        ].apply(clean_text)
    else:
        result["Party Name"] = ""

    # Taxable

    if detected["Taxable"]:
        result["Taxable Value"] = df[
            detected["Taxable"]
        ].apply(clean_amount)
    else:
        result["Taxable Value"] = 0.0

    # IGST

    if detected["IGST"]:
        result["IGST"] = df[
            detected["IGST"]
        ].apply(clean_amount)
    else:
        result["IGST"] = 0.0

    # CGST

    if detected["CGST"]:
        result["CGST"] = df[
            detected["CGST"]
        ].apply(clean_amount)
    else:
        result["CGST"] = 0.0

    # SGST

    if detected["SGST"]:
        result["SGST"] = df[
            detected["SGST"]
        ].apply(clean_amount)
    else:
        result["SGST"] = 0.0

    # Invoice Value

    if detected["InvoiceValue"]:

        result["Invoice Value"] = df[
            detected["InvoiceValue"]
        ].apply(clean_amount)

    else:

        result["Invoice Value"] = (
            result["Taxable Value"]
            + result["IGST"]
            + result["CGST"]
            + result["SGST"]
        )

    result["Source Sheet"] = (
        df["_SOURCE_SHEET"]
        if "_SOURCE_SHEET" in df.columns
        else ""
    )

    total_rows = len(result)

    blank_mask = (
        (result["GSTIN"] == "")
        |
        (result["Invoice Number"] == "")
    )

    dropped_rows = int(
        blank_mask.sum()
    )

    cleaned = result[
        ~blank_mask
    ].copy()

    dup_mask = cleaned.duplicated(
        subset=[
            "GSTIN",
            "Invoice Number"
        ],
        keep=False
    )

    duplicate_count = int(
        dup_mask.sum()
    )

    invalid_gstin_count = int(
        cleaned["GSTIN"]
        .apply(
            lambda g: not is_valid_gstin(g)
        )
        .sum()
    )

    quality = {
        "total_rows": total_rows,
        "dropped_rows": dropped_rows,
        "duplicate_rows": duplicate_count,
        "invalid_gstin_rows": invalid_gstin_count,
        "detected_columns": detected,
    }

    return (
        cleaned.reset_index(drop=True),
        quality
    )


# =========================================================
# NEW PORTAL NOTE / RCM ANALYSIS
# =========================================================

def find_note_type_column(df):
    possible = [
        "Note Type",
        "Document Type",
        "Debit/Credit Note",
        "Debit Credit Note",
        "Type",
        "Note Type (D/C)"
    ]

    return find_column(
        df,
        possible
    )


def classify_notes(cdn_raw):
    """
    Converts portal CDN data into separate Debit Note
    and Credit Note dataframes.
    """

    if cdn_raw is None or cdn_raw.empty:
        return pd.DataFrame(), pd.DataFrame()

    note_col = find_note_type_column(cdn_raw)

    if not note_col:
        return (
            cdn_raw.copy(),
            pd.DataFrame()
        )

    values = (
        cdn_raw[note_col]
        .fillna("")
        .astype(str)
        .str.upper()
    )

    debit_mask = (
        values.str.contains("DEBIT", na=False)
        |
        values.str.contains(r"\bDN\b", regex=True, na=False)
    )

    credit_mask = (
        values.str.contains("CREDIT", na=False)
        |
        values.str.contains(r"\bCN\b", regex=True, na=False)
    )

    debit_df = cdn_raw[
        debit_mask
    ].copy()

    credit_df = cdn_raw[
        credit_mask
    ].copy()

    return debit_df, credit_df


def find_rcm_column(df):
    possible = [
        "Reverse Charge",
        "Reverse charge",
        "RCM",
        "Reverse Charge Flag",
        "Supply Attract Reverse Charge",
        "Is Reverse Charge"
    ]

    return find_column(
        df,
        possible
    )


def analyse_rcm(two_b_raw, rcm_raw=None):
    """
    RCM can appear inside B2B itself as Reverse Charge = Yes.
    Separate RCM sheet is also considered if present.
    """

    rcm_b2b = pd.DataFrame()
    rcm_sheet = pd.DataFrame()

    if two_b_raw is not None and not two_b_raw.empty:

        rcm_col = find_rcm_column(two_b_raw)

        if rcm_col:

            values = (
                two_b_raw[rcm_col]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.strip()
            )

            rcm_mask = values.isin(
                [
                    "YES",
                    "Y",
                    "TRUE",
                    "1"
                ]
            )

            rcm_b2b = two_b_raw[
                rcm_mask
            ].copy()

    if rcm_raw is not None and not rcm_raw.empty:
        rcm_sheet = rcm_raw.copy()

    if not rcm_b2b.empty and not rcm_sheet.empty:

        combined = pd.concat(
            [rcm_b2b, rcm_sheet],
            ignore_index=True
        )

    elif not rcm_b2b.empty:

        combined = rcm_b2b

    else:

        combined = rcm_sheet

    return combined


def portal_document_stats(
    two_b_raw,
    cdn_raw,
    rcm_raw
):

    debit_df, credit_df = classify_notes(
        cdn_raw
    )

    rcm_df = analyse_rcm(
        two_b_raw,
        rcm_raw
    )

    return {
        "debit_notes": debit_df,
        "credit_notes": credit_df,
        "rcm": rcm_df,
        "debit_count": len(debit_df),
        "credit_count": len(credit_df),
        "rcm_count": len(rcm_df),
    }


# =========================================================
# RECONCILIATION
# =========================================================

def build_dedup_key(df):

    df = df.copy()

    base_key = (
        df["GSTIN"]
        + "|"
        + df["Invoice Number"]
    )

    occurrence = (
        base_key.groupby(base_key)
        .cumcount()
    )

    df["KEY"] = (
        base_key
        + "|"
        + occurrence.astype(str)
    )

    return df


def suggest_close_matches(
    missing_row,
    other_df,
    field="Invoice Number",
    cutoff=0.82
):

    same_gstin = other_df[
        other_df["GSTIN"]
        == missing_row["GSTIN"]
    ]

    if same_gstin.empty:
        return None

    candidates = same_gstin[
        field
    ].tolist()

    best = difflib.get_close_matches(
        missing_row[field],
        candidates,
        n=1,
        cutoff=cutoff
    )

    if best:

        match_row = same_gstin[
            same_gstin[field]
            == best[0]
        ].iloc[0]

        return (
            best[0],
            match_row.get(
                "Invoice Value",
                None
            )
        )

    return None


def reconcile(
    two_b,
    books,
    tolerance,
    enable_fuzzy=True
):

    two_b_k = build_dedup_key(
        two_b
    )

    books_k = build_dedup_key(
        books
    )

    result = pd.merge(
        books_k,
        two_b_k,
        on="KEY",
        how="outer",
        suffixes=(
            "_Books",
            "_2B"
        ),
        indicator=True
    )

    statuses = []
    differences = []
    suggestions = []

    for _, row in result.iterrows():

        if row["_merge"] == "left_only":

            statuses.append(
                "Missing in 2B"
            )

            differences.append(
                row["IGST_Books"]
                + row["CGST_Books"]
                + row["SGST_Books"]
            )

            suggestion = ""

            if enable_fuzzy:

                probe = {
                    "GSTIN": row[
                        "GSTIN_Books"
                    ],
                    "Invoice Number": row[
                        "Invoice Number_Books"
                    ]
                }

                found = suggest_close_matches(
                    probe,
                    two_b
                )

                if found:

                    suggestion = (
                        f"Possible match in 2B: "
                        f"'{found[0]}' — "
                        f"check formatting/typo"
                    )

            suggestions.append(
                suggestion
            )

        elif row["_merge"] == "right_only":

            statuses.append(
                "Missing in Books"
            )

            differences.append(
                row["IGST_2B"]
                + row["CGST_2B"]
                + row["SGST_2B"]
            )

            suggestion = ""

            if enable_fuzzy:

                probe = {
                    "GSTIN": row[
                        "GSTIN_2B"
                    ],
                    "Invoice Number": row[
                        "Invoice Number_2B"
                    ]
                }

                found = suggest_close_matches(
                    probe,
                    books
                )

                if found:

                    suggestion = (
                        f"Possible match in Books: "
                        f"'{found[0]}' — "
                        f"check formatting/typo"
                    )

            suggestions.append(
                suggestion
            )

        else:

            taxable_diff = abs(
                row["Taxable Value_Books"]
                -
                row["Taxable Value_2B"]
            )

            igst_diff = abs(
                row["IGST_Books"]
                -
                row["IGST_2B"]
            )

            cgst_diff = abs(
                row["CGST_Books"]
                -
                row["CGST_2B"]
            )

            sgst_diff = abs(
                row["SGST_Books"]
                -
                row["SGST_2B"]
            )

            invoice_diff = abs(
                row["Invoice Value_Books"]
                -
                row["Invoice Value_2B"]
            )

            total_diff = (
                igst_diff
                + cgst_diff
                + sgst_diff
            )

            if (
                taxable_diff <= tolerance
                and igst_diff <= tolerance
                and cgst_diff <= tolerance
                and sgst_diff <= tolerance
                and invoice_diff <= tolerance
            ):

                statuses.append(
                    "Matched"
                )

            else:

                statuses.append(
                    "Value Mismatch"
                )

            differences.append(
                total_diff
            )

            suggestions.append("")

    result["Status"] = statuses

    result["ITC Difference"] = differences

    result["Match Suggestion"] = suggestions

    return result


# =========================================================
# DISPLAY DATA
# =========================================================

def prepare_display(df):

    output = pd.DataFrame()

    def get_column(name):

        if name in df.columns:
            return df[name]

        return pd.Series(
            [""] * len(df),
            index=df.index
        )

    gstin_books = (
        get_column("GSTIN_Books")
        .fillna("")
    )

    gstin_2b = (
        get_column("GSTIN_2B")
        .fillna("")
    )

    output["GSTIN"] = gstin_books.where(
        gstin_books != "",
        gstin_2b
    )

    party_books = (
        get_column("Party Name_Books")
        .fillna("")
    )

    party_2b = (
        get_column("Party Name_2B")
        .fillna("")
    )

    output["Party Name"] = party_books.where(
        party_books != "",
        party_2b
    )

    inv_books = (
        get_column("Invoice Number_Books")
        .fillna("")
    )

    inv_2b = (
        get_column("Invoice Number_2B")
        .fillna("")
    )

    output["Invoice Number"] = inv_books.where(
        inv_books != "",
        inv_2b
    )

    output["Invoice Date"] = (
        get_column("Invoice Date_Books")
        .where(
            get_column(
                "Invoice Date_Books"
            ).notna(),
            get_column(
                "Invoice Date_2B"
            )
        )
    )

    output["Taxable - Books"] = get_column(
        "Taxable Value_Books"
    )

    output["Taxable - 2B"] = get_column(
        "Taxable Value_2B"
    )

    output["IGST - Books"] = get_column(
        "IGST_Books"
    )

    output["IGST - 2B"] = get_column(
        "IGST_2B"
    )

    output["CGST - Books"] = get_column(
        "CGST_Books"
    )

    output["CGST - 2B"] = get_column(
        "CGST_2B"
    )

    output["SGST - Books"] = get_column(
        "SGST_Books"
    )

    output["SGST - 2B"] = get_column(
        "SGST_2B"
    )

    output["Invoice Value - Books"] = get_column(
        "Invoice Value_Books"
    )

    output["Invoice Value - 2B"] = get_column(
        "Invoice Value_2B"
    )

    output["ITC Difference"] = get_column(
        "ITC Difference"
    )

    output["Status"] = get_column(
        "Status"
    )

    output["Match Suggestion"] = get_column(
        "Match Suggestion"
    )

    return output.reset_index(
        drop=True
    )


def apply_match_overrides(display_df):

    overrides = st.session_state.get(
        "match_suggestion_overrides",
        {}
    )

    if not overrides or display_df.empty:
        return display_df

    keys = (
        display_df["GSTIN"]
        + "|"
        + display_df["Invoice Number"]
    )

    display_df = display_df.copy()

    display_df["Match Suggestion"] = [
        overrides.get(
            k,
            v
        )
        for k, v in zip(
            keys,
            display_df["Match Suggestion"]
        )
    ]

    return display_df


def vendor_summary(display_df):

    grouped = display_df.groupby(
        [
            "GSTIN",
            "Party Name"
        ],
        dropna=False
    ).agg(
        Total_Invoices=(
            "Invoice Number",
            "count"
        ),

        Matched=(
            "Status",
            lambda s:
            (s == "Matched").sum()
        ),

        Missing_in_2B=(
            "Status",
            lambda s:
            (s == "Missing in 2B").sum()
        ),

        Missing_in_Books=(
            "Status",
            lambda s:
            (s == "Missing in Books").sum()
        ),

        Value_Mismatch=(
            "Status",
            lambda s:
            (s == "Value Mismatch").sum()
        ),

        ITC_at_Risk=(
            "ITC Difference",
            "sum"
        ),
    ).reset_index()

    grouped = grouped.sort_values(
        "ITC_at_Risk",
        ascending=False
    )

    grouped["ITC_at_Risk"] = (
        grouped["ITC_at_Risk"]
        .round(2)
    )

    return grouped


# =========================================================
# EXCEL REPORT
# =========================================================

def _style_worksheet_header(
    ws,
    ncols
):

    header_fill = PatternFill(
        start_color="0F2A5F",
        end_color="0F2A5F",
        fill_type="solid"
    )

    header_font = Font(
        name="Arial",
        bold=True,
        color="FFFFFF",
        size=10
    )

    thin = Side(
        style="thin",
        color="D9D9D9"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for col in range(
        1,
        ncols + 1
    ):

        c = ws.cell(
            row=1,
            column=col
        )

        c.fill = header_fill

        c.font = header_font

        c.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        c.border = border

    ws.freeze_panes = "A2"


def _autofit(
    ws,
    max_width=45
):

    for col_cells in ws.columns:

        length = max(
            (
                len(str(c.value))
                if c.value is not None
                else 0
            )
            for c in col_cells
        )

        col_letter = get_column_letter(
            col_cells[0].column
        )

        ws.column_dimensions[
            col_letter
        ].width = min(
            max(length + 3, 10),
            max_width
        )


def _format_data_rows(
    ws,
    ncols,
    status_col_index=None
):

    thin = Side(
        style="thin",
        color="EDEDED"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    money_cols = set()

    for idx, cell in enumerate(
        ws[1],
        start=1
    ):

        header = str(
            cell.value or ""
        )

        if any(
            tag in header
            for tag in [
                "Taxable",
                "IGST",
                "CGST",
                "SGST",
                "Invoice Value",
                "ITC"
            ]
        ):

            money_cols.add(idx)

    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        max_col=ncols
    ):

        for cell in row:

            cell.font = Font(
                name="Arial",
                size=10
            )

            cell.border = border

            if cell.column in money_cols:

                cell.number_format = (
                    "#,##0.00"
                )

        if status_col_index:

            status_val = row[
                status_col_index - 1
            ].value

            color = STATUS_COLORS.get(
                status_val
            )

            if color:

                fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type="solid"
                )

                font = Font(
                    name="Arial",
                    size=9,
                    bold=True,
                    color="FFFFFF"
                )

                cell = row[
                    status_col_index - 1
                ]

                cell.fill = fill

                cell.font = font

                cell.alignment = Alignment(
                    horizontal="center"
                )


def create_excel(
    result,
    quality_2b,
    quality_books,
    tolerance
):

    display = prepare_display(
        result
    )

    display = apply_match_overrides(
        display
    )

    v_summary = vendor_summary(
        display
    )

    total = len(display)

    matched = int(
        (
            display["Status"]
            == "Matched"
        ).sum()
    )

    missing_2b = int(
        (
            display["Status"]
            == "Missing in 2B"
        ).sum()
    )

    missing_books = int(
        (
            display["Status"]
            == "Missing in Books"
        ).sum()
    )

    mismatch = int(
        (
            display["Status"]
            == "Value Mismatch"
        ).sum()
    )

    itc_at_risk = display.loc[
        display["Status"] != "Matched",
        "ITC Difference"
    ].sum()

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        summary_df = pd.DataFrame({

            "Metric": [
                "Total Invoices Analysed",
                "Matched",
                "Missing in 2B",
                "Missing in Books",
                "Value Mismatch",
                "Match Rate (%)",
                "Total ITC at Risk (Rs.)",
                "Mismatch Tolerance Used (Rs.)",
            ],

            "Value": [
                total,
                matched,
                missing_2b,
                missing_books,
                mismatch,
                round(
                    (
                        matched
                        / total
                        * 100
                    ),
                    2
                )
                if total
                else 0,

                round(
                    itc_at_risk,
                    2
                ),

                tolerance,
            ]
        })

        summary_df.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

        display_export = display.copy()

        if "Invoice Date" in display_export.columns:

            display_export[
                "Invoice Date"
            ] = pd.to_datetime(
                display_export[
                    "Invoice Date"
                ],
                errors="coerce"
            ).dt.strftime(
                "%d-%m-%Y"
            )

        display_export.to_excel(
            writer,
            sheet_name="Complete Reconciliation",
            index=False
        )

        for status in [
            "Matched",
            "Missing in 2B",
            "Missing in Books",
            "Value Mismatch"
        ]:

            temp = display_export[
                display_export["Status"]
                == status
            ]

            temp.to_excel(
                writer,
                sheet_name=status[:31],
                index=False
            )

        v_summary.to_excel(
            writer,
            sheet_name="Vendor Summary",
            index=False
        )

    output.seek(0)

    wb = load_workbook(
        output
    )

    ws_summary = wb[
        "Summary"
    ]

    _style_worksheet_header(
        ws_summary,
        ws_summary.max_column
    )

    _format_data_rows(
        ws_summary,
        ws_summary.max_column
    )

    _autofit(
        ws_summary
    )

    main_sheet_names = [
        "Complete Reconciliation",
        "Matched",
        "Missing in 2B",
        "Missing in Books",
        "Value Mismatch"
    ]

    for name in main_sheet_names:

        ws = wb[name]

        if ws.max_row < 1:
            continue

        _style_worksheet_header(
            ws,
            ws.max_column
        )

        status_idx = None

        for i, cell in enumerate(
            ws[1],
            start=1
        ):

            if cell.value == "Status":

                status_idx = i

                break

        _format_data_rows(
            ws,
            ws.max_column,
            status_col_index=status_idx
        )

        _autofit(ws)

    ws_vendor = wb[
        "Vendor Summary"
    ]

    _style_worksheet_header(
        ws_vendor,
        ws_vendor.max_column
    )

    _format_data_rows(
        ws_vendor,
        ws_vendor.max_column
    )

    _autofit(
        ws_vendor
    )

    final_output = BytesIO()

    wb.save(
        final_output
    )

    final_output.seek(0)

    return final_output


# =========================================================
# HEADER
# =========================================================

st.markdown("""
<div class="hero">
    <div class="hero-topline">
        <div class="hero-kicker">✦ AI-powered reconciliation workspace</div>
        <div class="hero-live"><span class="hero-live-dot"></span> Ready to reconcile</div>
    </div>

    <div class="hero-main-title">
        <h1>GST Reconciliation Pro<span class="hero-cursor"></span></h1>
        <p>Turn messy GST 2B & Books data into a clean, actionable reconciliation — with smart matching, exception analysis and ITC risk visibility.</p>
        <div class="hero-pills">
            <span class="hero-pill">⚡ Smart Matching</span>
            <span class="hero-pill">🧠 Fuzzy Suggestions</span>
            <span class="hero-pill">📊 ITC Analytics</span>
            <span class="hero-pill">📥 Excel Report</span>
        </div>
    </div>

    <div class="ai-stage">
        <div class="ai-stage-glow"></div>

        <div class="ai-chip one">2B DATA</div>
        <div class="ai-chip two">AI MATCH</div>
        <div class="ai-chip three">ITC CHECK</div>
        <div class="ai-chip four">BOOKS</div>

        <div class="worker worker-yellow">
            <span class="worker-badge">Reader</span>
            <div class="worker-eyes"><span></span><span></span></div>
            <div class="worker-mouth"></div>
            <div class="worker-arm left-arm"></div>
            <div class="worker-arm right-arm"></div>
        </div>

        <div class="worker worker-pink">
            <span class="worker-badge">Matcher</span>
            <div class="worker-eyes"><span></span><span></span></div>
            <div class="worker-mouth"></div>
            <div class="worker-arm left-arm"></div>
            <div class="worker-arm right-arm"></div>
        </div>

        <div class="worker worker-blue">
            <span class="worker-badge">Analyzer</span>
            <div class="worker-eyes"><span></span><span></span></div>
            <div class="worker-mouth"></div>
            <div class="worker-arm left-arm"></div>
            <div class="worker-arm right-arm"></div>
        </div>

        <div class="worker worker-green">
            <span class="worker-badge">Auditor</span>
            <div class="worker-eyes"><span></span><span></span></div>
            <div class="worker-mouth"></div>
            <div class="worker-arm left-arm"></div>
            <div class="worker-arm right-arm"></div>
            <div class="worker-hand"></div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:

    st.header(
        "⚙️ Reconciliation Settings"
    )

    tolerance = st.number_input(
        "Mismatch Tolerance ₹",
        min_value=0.0,
        value=2.0,
        step=0.50,
        help=(
            "Differences within this amount "
            "are still treated as a Match."
        )
    )

    enable_fuzzy = st.checkbox(
        "🔍 Suggest possible matches for gaps",
        value=True,
        help=(
            "Flags likely formatting/typo differences "
            "instead of a genuine missing invoice."
        )
    )

    st.divider()

    st.subheader(
        "🔑 Matching Key"
    )

    st.info(
        "GSTIN + Invoice Number\n\n"
        "(duplicate invoices are paired in sequence)"
    )

    st.divider()

    st.subheader(
        "📌 Status Legend"
    )

    st.markdown(
        "🟢 **Matched** — tallies within tolerance"
    )

    st.markdown(
        "🔴 **Missing in 2B** — in Books, not filed by supplier"
    )

    st.markdown(
        "🟠 **Missing in Books** — filed by supplier, not booked"
    )

    st.markdown(
        "🟣 **Value Mismatch** — amounts differ beyond tolerance"
    )


# =========================================================
# UPLOAD
# =========================================================

st.markdown(
    '<div class="section-title">📂 Upload GST Data</div>',
    unsafe_allow_html=True
)

col1, col2 = st.columns(2)


with col1:

    st.subheader(
        "📥 GST 2B File"
    )

    st.caption(
        "Upload GST portal 2B Excel file. Multiple sheets supported."
    )

    two_b_file = st.file_uploader(
        "Choose GST 2B Excel",
        type=["xlsx", "xls"],
        key="gst_2b_upload"
    )


with col2:

    st.subheader(
        "📚 Books File"
    )

    st.caption(
        "Upload your purchase/books Excel file."
    )

    books_file = st.file_uploader(
        "Choose Books Excel",
        type=["xlsx", "xls"],
        key="books_upload"
    )


# =========================================================
# OVERRIDES
# =========================================================

overrides_2b = {}
overrides_books = {}


# =========================================================
# LOAD FILES
# =========================================================

if two_b_file and books_file:

    try:

        # -----------------------------------------------------
        # PORTAL 2B
        # -----------------------------------------------------

        if (
            "portal_b2b_raw"
            not in st.session_state
            or
            st.session_state.get(
                "two_b_name"
            )
            != two_b_file.name
        ):

            portal_data = read_portal_2b(
                two_b_file
            )

            st.session_state[
                "portal_b2b_raw"
            ] = portal_data[
                "b2b_raw"
            ]

            st.session_state[
                "portal_cdn_raw"
            ] = portal_data[
                "cdn_raw"
            ]

            st.session_state[
                "portal_rcm_raw"
            ] = portal_data[
                "rcm_raw"
            ]

            st.session_state[
                "two_b_sheets"
            ] = portal_data[
                "b2b_sheets"
            ]

            st.session_state[
                "portal_cdn_sheets"
            ] = portal_data[
                "cdn_sheets"
            ]

            st.session_state[
                "portal_rcm_sheets"
            ] = portal_data[
                "rcm_sheets"
            ]

            st.session_state[
                "two_b_skipped"
            ] = portal_data[
                "skipped_sheets"
            ]

            st.session_state[
                "portal_all_sheets"
            ] = portal_data[
                "all_sheets"
            ]

            st.session_state[
                "two_b_name"
            ] = two_b_file.name

            st.session_state[
                "match_suggestion_overrides"
            ] = {}

        # -----------------------------------------------------
        # BOOKS
        # -----------------------------------------------------

        if (
            "books_raw"
            not in st.session_state
            or
            st.session_state.get(
                "books_name"
            )
            != books_file.name
        ):

            raw, sheets, skipped = read_excel(
                books_file
            )

            st.session_state[
                "books_raw"
            ] = raw

            st.session_state[
                "books_sheets"
            ] = sheets

            st.session_state[
                "books_skipped"
            ] = skipped

            st.session_state[
                "books_name"
            ] = books_file.name

            st.session_state[
                "match_suggestion_overrides"
            ] = {}

        two_b_raw = st.session_state[
            "portal_b2b_raw"
        ]

        books_raw = st.session_state[
            "books_raw"
        ]

        # =====================================================
        # COLUMN MAPPING
        # =====================================================

        with st.expander(
            "🔧 Column Mapping (verify or override auto-detection)"
        ):

            st.caption(
                "Har alag Excel export column names alag rakhta hai — "
                "humne best-guess mapping kar di hai. "
                "Zaroorat ho to yahan se change karein."
            )

            m1, m2 = st.columns(2)

            # -------------------------------------------------
            # 2B
            # -------------------------------------------------

            with m1:

                st.markdown(
                    "**GST 2B B2B columns**"
                )

                auto_2b = auto_detect(
                    two_b_raw
                )

                options_2b = [
                    "— None —"
                ] + list(
                    two_b_raw.columns
                )

                for field in (
                    REQUIRED_FIELDS
                    + OPTIONAL_FIELDS
                ):

                    default_val = (
                        auto_2b.get(field)
                        or
                        "— None —"
                    )

                    default_idx = (
                        options_2b.index(
                            default_val
                        )
                        if
                        default_val
                        in options_2b
                        else 0
                    )

                    overrides_2b[field] = st.selectbox(
                        FIELD_LABELS[field],
                        options_2b,
                        index=default_idx,
                        key=f"2b_{field}"
                    )

            # -------------------------------------------------
            # BOOKS
            # -------------------------------------------------

            with m2:

                st.markdown(
                    "**Books columns**"
                )

                auto_books = auto_detect(
                    books_raw
                )

                options_books = [
                    "— None —"
                ] + list(
                    books_raw.columns
                )

                for field in (
                    REQUIRED_FIELDS
                    + OPTIONAL_FIELDS
                ):

                    default_val = (
                        auto_books.get(field)
                        or
                        "— None —"
                    )

                    default_idx = (
                        options_books.index(
                            default_val
                        )
                        if
                        default_val
                        in options_books
                        else 0
                    )

                    overrides_books[field] = st.selectbox(
                        FIELD_LABELS[field],
                        options_books,
                        index=default_idx,
                        key=f"bk_{field}"
                    )

    except Exception as error:

        st.error(
            "❌ File padhne mein error aayi"
        )

        st.error(
            str(error)
        )


# =========================================================
# PROCESS
# =========================================================

if (
    two_b_file
    and books_file
    and "portal_b2b_raw"
    in st.session_state
):

    st.write("")

    if st.button(
        "🚀 START GST RECONCILIATION",
        type="primary",
        use_container_width=True
    ):

        try:

            mascot_placeholder = st.empty()

            mascot_placeholder.markdown("""
            <div class="run-overlay">
                <div class="worker-loader">
                    <div class="loader-worker loader-yellow"></div>
                    <div class="loader-worker loader-blue"></div>
                    <div class="loader-worker loader-green"></div>
                </div>
                <p class="loading-text">
                    Matching invoices, GSTIN by GSTIN
                    <span class="loading-dots"><span>•</span><span>•</span><span>•</span></span>
                </p>
            </div>
            """, unsafe_allow_html=True)

            with st.spinner(
                "Analysing GST 2B and Books..."
            ):

                # -------------------------------------------------
                # IMPORTANT:
                # Only B2B is used for normal invoice reconciliation.
                # CDN and RCM remain separate.
                # -------------------------------------------------

                two_b, quality_2b = standardize(
                    st.session_state[
                        "portal_b2b_raw"
                    ],
                    overrides_2b
                )

                books, quality_books = standardize(
                    st.session_state[
                        "books_raw"
                    ],
                    overrides_books
                )

                if two_b.empty:

                    raise ValueError(
                        "GST 2B B2B sheet mein valid invoice data nahi mila."
                    )

                if books.empty:

                    raise ValueError(
                        "Books file mein valid invoice data nahi mila."
                    )

                result = reconcile(
                    two_b,
                    books,
                    tolerance,
                    enable_fuzzy=enable_fuzzy
                )

                # -------------------------------------------------
                # PORTAL DOCUMENT ANALYSIS
                # -------------------------------------------------

                portal_stats = portal_document_stats(
                    st.session_state.get(
                        "portal_b2b_raw",
                        pd.DataFrame()
                    ),
                    st.session_state.get(
                        "portal_cdn_raw",
                        pd.DataFrame()
                    ),
                    st.session_state.get(
                        "portal_rcm_raw",
                        pd.DataFrame()
                    )
                )

                st.session_state[
                    "result"
                ] = result

                st.session_state[
                    "two_b"
                ] = two_b

                st.session_state[
                    "books"
                ] = books

                st.session_state[
                    "quality_2b"
                ] = quality_2b

                st.session_state[
                    "quality_books"
                ] = quality_books

                st.session_state[
                    "tolerance_used"
                ] = tolerance

                st.session_state[
                    "selected_status"
                ] = "Missing in 2B"

                st.session_state[
                    "portal_stats"
                ] = portal_stats

                st.session_state.setdefault(
                    "match_suggestion_overrides",
                    {}
                )

            mascot_placeholder.empty()

            st.success(
                "✅ GST Reconciliation completed successfully."
            )

        except Exception as error:

            mascot_placeholder.empty()

            st.error(
                "❌ File processing error"
            )

            st.error(
                str(error)
            )


# =========================================================
# DATA QUALITY PANEL
# =========================================================

if (
    "quality_2b"
    in st.session_state
    and
    "result"
    in st.session_state
):

    q2b = st.session_state[
        "quality_2b"
    ]

    qbk = st.session_state[
        "quality_books"
    ]

    quality_flags = []

    if (
        q2b["dropped_rows"]
        or
        qbk["dropped_rows"]
    ):

        quality_flags.append(
            f"⚠️ Rows skipped due to blank GSTIN/Invoice — "
            f"2B: {q2b['dropped_rows']}, "
            f"Books: {qbk['dropped_rows']}"
        )

    if (
        q2b["duplicate_rows"]
        or
        qbk["duplicate_rows"]
    ):

        quality_flags.append(
            f"⚠️ Duplicate GSTIN+Invoice rows detected — "
            f"2B: {q2b['duplicate_rows']}, "
            f"Books: {qbk['duplicate_rows']} "
            "(paired in sequence during matching)"
        )

    if (
        q2b["invalid_gstin_rows"]
        or
        qbk["invalid_gstin_rows"]
    ):

        quality_flags.append(
            f"⚠️ Rows with non-standard GSTIN format — "
            f"2B: {q2b['invalid_gstin_rows']}, "
            f"Books: {qbk['invalid_gstin_rows']}"
        )

    with st.expander(
        "🩺 Data Quality Checks",
        expanded=bool(quality_flags)
    ):

        if quality_flags:

            for flag in quality_flags:

                st.markdown(
                    f'<div class="warn-box">{flag}</div>',
                    unsafe_allow_html=True
                )

        else:

            st.success(
                "Koi data quality issue nahi mila — files clean hain."
            )


# =========================================================
# DASHBOARD
# =========================================================

if "result" in st.session_state:

    result = st.session_state[
        "result"
    ]

    two_b = st.session_state[
        "two_b"
    ]

    books = st.session_state[
        "books"
    ]

    total = len(result)

    matched = int(
        (
            result["Status"]
            == "Matched"
        ).sum()
    )

    missing_2b = int(
        (
            result["Status"]
            == "Missing in 2B"
        ).sum()
    )

    missing_books = int(
        (
            result["Status"]
            == "Missing in Books"
        ).sum()
    )

    mismatch = int(
        (
            result["Status"]
            == "Value Mismatch"
        ).sum()
    )

    match_rate = round(
        (
            matched
            / total
            * 100
        ),
        1
    ) if total else 0.0

    # ---------------------------------------------------------
    # PORTAL STATS
    # ---------------------------------------------------------

    portal_stats = st.session_state.get(
        "portal_stats",
        {
            "debit_count": 0,
            "credit_count": 0,
            "rcm_count": 0
        }
    )

    debit_count = portal_stats.get(
        "debit_count",
        0
    )

    credit_count = portal_stats.get(
        "credit_count",
        0
    )

    rcm_count = portal_stats.get(
        "rcm_count",
        0
    )

    # ========================================================
    # DASHBOARD
    # ========================================================

    st.markdown(
        '<div class="section-title">📊 Reconciliation Dashboard</div>',
        unsafe_allow_html=True
    )

    c1, c2, c3, c4, c5, c6 = st.columns(6)

    with c1:

        st.markdown(
            f"""
            <div class="kpi blue">
                <div class="kpi-title">
                    TOTAL INVOICES
                </div>
                <div class="kpi-value">
                    {total:,}
                </div>
                <div class="kpi-desc">
                    All analysed B2B records
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c2:

        st.markdown(
            f"""
            <div class="kpi green">
                <div class="kpi-title">
                    MATCHED
                </div>
                <div class="kpi-value">
                    {matched:,}
                </div>
                <div class="kpi-desc">
                    {match_rate}% match rate
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c3:

        st.markdown(
            f"""
            <div class="kpi red">
                <div class="kpi-title">
                    MISSING IN 2B
                </div>
                <div class="kpi-value">
                    {missing_2b:,}
                </div>
                <div class="kpi-desc">
                    Books → not in 2B
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c4:

        st.markdown(
            f"""
            <div class="kpi orange">
                <div class="kpi-title">
                    MISSING IN BOOKS
                </div>
                <div class="kpi-value">
                    {missing_books:,}
                </div>
                <div class="kpi-desc">
                    2B → not in Books
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c5:

        st.markdown(
            f"""
            <div class="kpi purple">
                <div class="kpi-title">
                    VALUE MISMATCH
                </div>
                <div class="kpi-value">
                    {mismatch:,}
                </div>
                <div class="kpi-desc">
                    Tax/value difference
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c6:

        itc_at_risk = result.loc[
            result["Status"] != "Matched",
            "ITC Difference"
        ].sum()

        st.markdown(
            f"""
            <div class="kpi teal">
                <div class="kpi-title">
                    ITC AT RISK
                </div>
                <div class="kpi-value">
                    ₹{itc_at_risk:,.0f}
                </div>
                <div class="kpi-desc">
                    Sum of open differences
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # ========================================================
    # NEW PORTAL DOCUMENT SUMMARY
    # ========================================================

    st.markdown(
        '<div class="section-title">📑 GST Portal Document Analysis</div>',
        unsafe_allow_html=True
    )

    d1, d2, d3 = st.columns(3)

    with d1:

        st.markdown(
            f"""
            <div class="kpi orange">
                <div class="kpi-title">
                    DEBIT NOTES
                </div>
                <div class="kpi-value">
                    {debit_count:,}
                </div>
                <div class="kpi-desc">
                    GST portal CDN records
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with d2:

        st.markdown(
            f"""
            <div class="kpi red">
                <div class="kpi-title">
                    CREDIT NOTES
                </div>
                <div class="kpi-value">
                    {credit_count:,}
                </div>
                <div class="kpi-desc">
                    GST portal CDN records
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with d3:

        st.markdown(
            f"""
            <div class="kpi purple">
                <div class="kpi-title">
                    RCM CASES
                </div>
                <div class="kpi-value">
                    {rcm_count:,}
                </div>
                <div class="kpi-desc">
                    Reverse Charge = Yes
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )


    # ========================================================
    # PORTAL DOCUMENT DETAILS
    # ========================================================

    p1, p2, p3 = st.columns(3)

    with p1:

        with st.expander(
            f"🟠 Debit Notes ({debit_count})"
        ):

            debit_df = portal_stats.get(
                "debit_notes",
                pd.DataFrame()
            )

            if not debit_df.empty:

                st.dataframe(
                    debit_df,
                    use_container_width=True,
                    hide_index=True
                )

            else:

                st.info(
                    "GST portal file mein Debit Note record nahi mila."
                )

    with p2:

        with st.expander(
            f"🔴 Credit Notes ({credit_count})"
        ):

            credit_df = portal_stats.get(
                "credit_notes",
                pd.DataFrame()
            )

            if not credit_df.empty:

                st.dataframe(
                    credit_df,
                    use_container_width=True,
                    hide_index=True
                )

            else:

                st.info(
                    "GST portal file mein Credit Note record nahi mila."
                )

    with p3:

        with st.expander(
            f"🟣 RCM Cases ({rcm_count})"
        ):

            rcm_df = portal_stats.get(
                "rcm",
                pd.DataFrame()
            )

            if not rcm_df.empty:

                st.dataframe(
                    rcm_df,
                    use_container_width=True,
                    hide_index=True
                )

            else:

                st.info(
                    "B2B mein Reverse Charge = Yes wala record nahi mila."
                )


    # ========================================================
    # INVOICE ANALYSIS
    # ========================================================

    st.markdown(
        '<div class="section-title">🔎 Invoice Analysis</div>',
        unsafe_allow_html=True
    )

    b1, b2, b3, b4 = st.columns(4)

    with b1:

        if st.button(
            f"🔴 Missing in 2B • {missing_2b}",
            use_container_width=True
        ):

            st.session_state[
                "selected_status"
            ] = "Missing in 2B"

    with b2:

        if st.button(
            f"🟠 Missing in Books • {missing_books}",
            use_container_width=True
        ):

            st.session_state[
                "selected_status"
            ] = "Missing in Books"

    with b3:

        if st.button(
            f"🟣 Value Mismatch • {mismatch}",
            use_container_width=True
        ):

            st.session_state[
                "selected_status"
            ] = "Value Mismatch"

    with b4:

        if st.button(
            f"🟢 Matched • {matched}",
            use_container_width=True
        ):

            st.session_state[
                "selected_status"
            ] = "Matched"


    selected_status = st.session_state.get(
        "selected_status",
        "Missing in 2B"
    )

    detail = result[
        result["Status"]
        == selected_status
    ].copy()

    st.subheader(
        f"📋 {selected_status} — {len(detail):,} Invoices"
    )

    explanations = {

        "Missing in 2B":
            "Books mein invoice hai, lekin GST 2B mein nahi mila.",

        "Missing in Books":
            "GST 2B mein invoice hai, lekin Books mein nahi mila.",

        "Value Mismatch":
            "GSTIN aur invoice number match hain, lekin taxable/tax/invoice value different hai.",

        "Matched":
            "Books aur GST 2B mein invoice successfully match hua.",
    }

    st.info(
        explanations[selected_status]
    )

    display = prepare_display(
        detail
    )

    display = apply_match_overrides(
        display
    )

    search_term = st.text_input(
        "🔍 Search within this list (GSTIN, Invoice No, or Party Name)",
        placeholder="e.g. 27AABCU9603R1ZM or INV/2025/1001"
    )

    if not display.empty:

        display["Invoice Date"] = pd.to_datetime(
            display["Invoice Date"],
            errors="coerce"
        ).dt.strftime(
            "%d-%m-%Y"
        )

        display["Invoice Date"] = (
            display["Invoice Date"]
            .fillna("-")
        )

        if search_term.strip():

            term = (
                search_term
                .strip()
                .lower()
            )

            mask = (
                display["GSTIN"]
                .str.lower()
                .str.contains(
                    term,
                    na=False
                )
                |
                display["Invoice Number"]
                .str.lower()
                .str.contains(
                    term,
                    na=False
                )
                |
                display["Party Name"]
                .str.lower()
                .str.contains(
                    term,
                    na=False
                )
            )

            display = display[
                mask
            ]

        st.caption(
            "✏️ 'Match Suggestion' column mein khud bhi type kar sakte ho — remarks ya manual match note."
        )

        edited = st.data_editor(
            display,
            use_container_width=True,
            hide_index=True,
            disabled=[
                c
                for c in display.columns
                if c != "Match Suggestion"
            ],
            column_config={
                "Match Suggestion":
                    st.column_config.TextColumn(
                        "Match Suggestion / Remarks",
                        help=(
                            "Apna remark ya manual match note yahan type karein"
                        ),
                        width="large",
                    )
            },
            key=f"editor_{selected_status}",
        )

        if not edited.empty:

            overrides = st.session_state.setdefault(
                "match_suggestion_overrides",
                {}
            )

            edit_keys = (
                edited["GSTIN"]
                + "|"
                + edited["Invoice Number"]
            )

            for k, v in zip(
                edit_keys,
                edited["Match Suggestion"]
            ):

                overrides[k] = v

        update_col, _ = st.columns(
            [1, 3]
        )

        with update_col:

            if st.button(
                "💾 Update Changes",
                key=f"save_{selected_status}",
                use_container_width=True
            ):

                overrides = st.session_state.setdefault(
                    "match_suggestion_overrides",
                    {}
                )

                edit_keys = (
                    edited["GSTIN"]
                    + "|"
                    + edited["Invoice Number"]
                )

                for k, v in zip(
                    edit_keys,
                    edited["Match Suggestion"]
                ):

                    overrides[k] = v

                st.success(
                    "✅ Changes saved. Ye notes ab tab tak bane rahenge jab tak koi nayi file upload na ho — dobara reconcile karne se bhi nahi udenge."
                )

        csv_bytes = (
            edited
            .to_csv(index=False)
            .encode("utf-8-sig")
        )

        st.download_button(
            f"⬇️ Download '{selected_status}' as CSV",
            data=csv_bytes,
            file_name=(
                f"GST_{selected_status.replace(' ', '_')}.csv"
            ),
            mime="text/csv",
        )

    else:

        st.success(
            "Is category mein koi invoice nahi hai."
        )


    # ========================================================
    # MANUAL SEARCH
    # ========================================================

    if (
        selected_status
        in (
            "Missing in 2B",
            "Missing in Books"
        )
        and
        not detail.empty
    ):

        other_df = (
            two_b
            if selected_status
            == "Missing in 2B"
            else books
        )

        other_label = (
            "GST 2B"
            if selected_status
            == "Missing in 2B"
            else "Books"
        )

        with st.expander(
            f"🔎 Manually search {other_label} for a possible match",
            expanded=False
        ):

            st.caption(
                f"Auto-suggestion ne kuch na pakda ho to yahan khud type karke {other_label} file mein dhoond lo — GSTIN, Invoice Number, ya Party Name, kisi se bhi search ho jayega."
            )

            manual_query = st.text_input(
                f"Type to search in {other_label}",
                key=f"manual_search_{selected_status}",
                placeholder="e.g. GSTIN ka part, invoice number, ya vendor ka naam"
            )

            if manual_query.strip():

                q = (
                    manual_query
                    .strip()
                    .lower()
                )

                other_disp = other_df.copy()

                match_mask = (
                    other_disp["GSTIN"]
                    .str.lower()
                    .str.contains(
                        q,
                        na=False
                    )
                    |
                    other_disp[
                        "Invoice Number"
                    ]
                    .str.lower()
                    .str.contains(
                        q,
                        na=False
                    )
                    |
                    other_disp[
                        "Party Name"
                    ]
                    .str.lower()
                    .str.contains(
                        q,
                        na=False
                    )
                )

                found_cols = [
                    "GSTIN",
                    "Party Name",
                    "Invoice Number",
                    "Invoice Date",
                    "Taxable Value",
                    "IGST",
                    "CGST",
                    "SGST",
                    "Invoice Value"
                ]

                found = other_disp.loc[
                    match_mask,
                    found_cols
                ].copy()

                if not found.empty:

                    found["Invoice Date"] = pd.to_datetime(
                        found["Invoice Date"],
                        errors="coerce"
                    ).dt.strftime(
                        "%d-%m-%Y"
                    )

                    found["Invoice Date"] = (
                        found["Invoice Date"]
                        .fillna("-")
                    )

                    st.dataframe(
                        found,
                        use_container_width=True,
                        hide_index=True
                    )

                    st.caption(
                        f"{len(found)} possible record(s) mile {other_label} mein."
                    )

                else:

                    st.warning(
                        f"Koi record nahi mila jo '{manual_query}' se match kare."
                    )

            else:

                st.caption(
                    "Search shuru karne ke liye upar type karo."
                )


    # ========================================================
    # VENDOR SUMMARY
    # ========================================================

    st.markdown(
        '<div class="section-title">🏢 Vendor-wise Summary</div>',
        unsafe_allow_html=True
    )

    st.caption(
        "Sabse zyada ITC-at-risk wale vendors sabse upar dikhaye gaye hain."
    )

    full_display = prepare_display(
        result
    )

    v_summary = vendor_summary(
        full_display
    )

    st.dataframe(
        v_summary.rename(
            columns={
                "Total_Invoices":
                    "Total Invoices",

                "Missing_in_2B":
                    "Missing in 2B",

                "Missing_in_Books":
                    "Missing in Books",

                "Value_Mismatch":
                    "Value Mismatch",

                "ITC_at_Risk":
                    "ITC at Risk (Rs.)",
            }
        ).style.format(
            {
                "ITC at Risk (Rs.)":
                    "₹{:,.2f}"
            }
        ),
        use_container_width=True,
        hide_index=True,
    )


    # ========================================================
    # ITC ANALYSIS
    # ========================================================

    st.markdown(
        '<div class="section-title">💰 ITC Analysis</div>',
        unsafe_allow_html=True
    )

    books_igst = books[
        "IGST"
    ].sum()

    books_cgst = books[
        "CGST"
    ].sum()

    books_sgst = books[
        "SGST"
    ].sum()

    two_b_igst = two_b[
        "IGST"
    ].sum()

    two_b_cgst = two_b[
        "CGST"
    ].sum()

    two_b_sgst = two_b[
        "SGST"
    ].sum()

    books_itc = (
        books_igst
        + books_cgst
        + books_sgst
    )

    two_b_itc = (
        two_b_igst
        + two_b_cgst
        + two_b_sgst
    )

    itc_difference = (
        books_itc
        - two_b_itc
    )

    i1, i2, i3 = st.columns(3)

    with i1:

        st.metric(
            "📚 Books Total ITC",
            f"₹{books_itc:,.2f}"
        )

    with i2:

        st.metric(
            "📥 GST 2B Total ITC",
            f"₹{two_b_itc:,.2f}"
        )

    with i3:

        st.metric(
            "⚖️ ITC Difference",
            f"₹{itc_difference:,.2f}"
        )

    tax_df = pd.DataFrame({

        "Tax Component": [
            "IGST",
            "CGST",
            "SGST",
            "TOTAL ITC"
        ],

        "Books": [
            books_igst,
            books_cgst,
            books_sgst,
            books_itc
        ],

        "GST 2B": [
            two_b_igst,
            two_b_cgst,
            two_b_sgst,
            two_b_itc
        ],

        "Difference": [
            books_igst - two_b_igst,
            books_cgst - two_b_cgst,
            books_sgst - two_b_sgst,
            itc_difference
        ]
    })

    st.dataframe(
        tax_df.style.format(
            {
                "Books":
                    "₹{:,.2f}",

                "GST 2B":
                    "₹{:,.2f}",

                "Difference":
                    "₹{:,.2f}"
            }
        ),
        use_container_width=True,
        hide_index=True
    )


    # ========================================================
    # CHARTS
    # ========================================================

    st.markdown(
        '<div class="section-title">📈 Visual Analysis</div>',
        unsafe_allow_html=True
    )

    chart1, chart2 = st.columns(2)

    with chart1:

        status_df = pd.DataFrame({

            "Status": [
                "Matched",
                "Missing in 2B",
                "Missing in Books",
                "Value Mismatch"
            ],

            "Count": [
                matched,
                missing_2b,
                missing_books,
                mismatch
            ]
        })

        status_df = status_df[
            status_df["Count"] > 0
        ]

        if not status_df.empty:

            fig = px.pie(
                status_df,
                names="Status",
                values="Count",
                hole=0.55,
                title="Reconciliation Status",
                color="Status",
                color_discrete_map={
                    "Matched":
                        "#16a34a",

                    "Missing in 2B":
                        "#dc2626",

                    "Missing in Books":
                        "#ea580c",

                    "Value Mismatch":
                        "#7c3aed",
                }
            )

            fig.update_traces(
                textinfo="percent+label"
            )

            st.plotly_chart(
                fig,
                use_container_width=True
            )


    with chart2:

        top_vendors = v_summary[
            v_summary["ITC_at_Risk"] > 0
        ].head(10)

        if not top_vendors.empty:

            fig3 = px.bar(
                top_vendors.sort_values(
                    "ITC_at_Risk"
                ),
                x="ITC_at_Risk",
                y="Party Name",
                orientation="h",
                title="Top 10 Vendors by ITC at Risk",
                labels={
                    "ITC_at_Risk":
                        "ITC at Risk (Rs.)",

                    "Party Name":
                        ""
                },
            )

            fig3.update_traces(
                marker_color="#7c3aed"
            )

            st.plotly_chart(
                fig3,
                use_container_width=True
            )

        else:

            itc_df = pd.DataFrame({

                "Tax": [
                    "IGST",
                    "CGST",
                    "SGST"
                ],

                "Amount": [
                    two_b_igst,
                    two_b_cgst,
                    two_b_sgst
                ]
            })

            itc_df = itc_df[
                itc_df["Amount"] > 0
            ]

            if not itc_df.empty:

                fig2 = px.pie(
                    itc_df,
                    names="Tax",
                    values="Amount",
                    hole=0.55,
                    title="GST 2B ITC Distribution"
                )

                fig2.update_traces(
                    textinfo="percent+label"
                )

                st.plotly_chart(
                    fig2,
                    use_container_width=True
                )


    # ========================================================
    # MONTHLY TREND
    # ========================================================

    combined_dates = pd.concat(
        [
            two_b[
                [
                    "Invoice Date",
                    "Taxable Value"
                ]
            ].assign(
                Source="GST 2B"
            ),

            books[
                [
                    "Invoice Date",
                    "Taxable Value"
                ]
            ].assign(
                Source="Books"
            ),
        ]
    )

    combined_dates = combined_dates.dropna(
        subset=[
            "Invoice Date"
        ]
    )

    if not combined_dates.empty:

        combined_dates["Month"] = (
            combined_dates[
                "Invoice Date"
            ]
            .dt.to_period("M")
            .astype(str)
        )

        monthly = (
            combined_dates
            .groupby(
                [
                    "Month",
                    "Source"
                ]
            )["Taxable Value"]
            .sum()
            .reset_index()
        )

        if (
            monthly["Month"]
            .nunique()
            > 1
        ):

            fig4 = px.bar(
                monthly,
                x="Month",
                y="Taxable Value",
                color="Source",
                barmode="group",
                title=(
                    "Month-wise Taxable Value: "
                    "GST 2B vs Books"
                ),
                color_discrete_map={
                    "GST 2B":
                        "#2563eb",

                    "Books":
                        "#16a34a"
                },
            )

            st.plotly_chart(
                fig4,
                use_container_width=True
            )


    # ========================================================
    # PORTAL SHEETS
    # ========================================================

    with st.expander(
        "📂 GST Portal Excel Sheets Automatically Analysed"
    ):

        st.info(
            "Reconciliation ke liye sirf B2B invoices use kiye gaye hain. "
            "Debit/Credit Notes aur RCM ko separately analyse kiya gaya hai."
        )

        col_a, col_b, col_c = st.columns(3)

        with col_a:

            st.subheader(
                "📥 B2B"
            )

            for sheet in st.session_state.get(
                "two_b_sheets",
                []
            ):

                st.write(
                    "•",
                    sheet
                )

        with col_b:

            st.subheader(
                "📝 Debit/Credit Notes"
            )

            for sheet in st.session_state.get(
                "portal_cdn_sheets",
                []
            ):

                st.write(
                    "•",
                    sheet
                )

        with col_c:

            st.subheader(
                "🔄 RCM"
            )

            for sheet in st.session_state.get(
                "portal_rcm_sheets",
                []
            ):

                st.write(
                    "•",
                    sheet
                )

        st.divider()

        st.caption(
            "Other portal sheets:"
        )

        all_sheets = st.session_state.get(
            "portal_all_sheets",
            []
        )

        used_sheets = (
            st.session_state.get(
                "two_b_sheets",
                []
            )
            +
            st.session_state.get(
                "portal_cdn_sheets",
                []
            )
            +
            st.session_state.get(
                "portal_rcm_sheets",
                []
            )
        )

        for sheet in all_sheets:

            if sheet not in used_sheets:

                st.write(
                    "•",
                    sheet
                )


    # ========================================================
    # EXPORT
    # ========================================================

    st.markdown(
        '<div class="section-title">📥 Export Report</div>',
        unsafe_allow_html=True
    )

    excel_report = create_excel(
        result,
        st.session_state[
            "quality_2b"
        ],
        st.session_state[
            "quality_books"
        ],
        st.session_state[
            "tolerance_used"
        ],
    )

    st.download_button(
        "📊 DOWNLOAD COMPLETE EXCEL REPORT",
        data=excel_report,
        file_name="GST_Reconciliation_Report.xlsx",
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )

    st.caption(
        "Report mein Summary, Complete Reconciliation, "
        "status-wise sheets, aur Vendor Summary — "
        "sab color-coded aur formatted hain."
    )


# =========================================================
# WELCOME
# =========================================================

else:

    st.markdown("""
    <div class="float-icons">
        <div class="float-icon">📥</div>
        <div class="float-icon">📚</div>
        <div class="float-icon">🧠</div>
        <div class="float-icon">🔍</div>
        <div class="float-icon">📊</div>
        <div class="float-icon">✅</div>
        <div class="float-icon">💰</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="info-box" style="animation: fadeInUp 0.6s ease;">

    <h3>👋 Welcome to GST Reconciliation Pro</h3>
    <div style="margin:-2px 0 14px; color:#2563eb; font-size:11px; font-weight:800; letter-spacing:.05em;">
        YOUR DIGITAL GST TEAM • READ → MATCH → ANALYSE → REPORT
    </div>

    Upload your <b>GST 2B</b> and <b>Books</b> Excel files above to get started.

    <br><br>

    The software automatically detects important GST columns even when the
    column names are different — and lets you fine-tune the mapping if needed.

    <br><br>

    <b>GST Portal Support:</b><br>
    B2B invoices are used for reconciliation while Debit Notes,
    Credit Notes and RCM cases are analysed separately.

    <br><br>

    <b>Matching Key:</b><br>
    GSTIN + Invoice Number (duplicates paired in sequence)

    <br><br>

    🔴 Missing in 2B<br>
    🟠 Missing in Books<br>
    🟣 Value Mismatch<br>
    🟢 Matched

    </div>
    """, unsafe_allow_html=True)

    st.markdown(
        '<div class="section-title">✨ What you get</div>',
        unsafe_allow_html=True
    )

    features = [
        (
            "🎯",
            "Smart Status Matching",
            "Matched, Missing in 2B, Missing in Books & Value Mismatch — auto classified."
        ),
        (
            "🧠",
            "Fuzzy Match Suggestions",
            "Typo or formatting mismatches get flagged instead of marked genuinely missing."
        ),
        (
            "🏢",
            "Vendor-wise ITC Risk",
            "See which vendors are holding up your ITC, ranked by risk amount."
        ),
        (
            "📊",
            "Polished Excel Export",
            "Color-coded, multi-sheet, ready-to-share report in one click."
        ),
    ]

    f_cols = st.columns(4)

    for f_col, (icon, title, desc) in zip(f_cols, features):

        with f_col:

            st.markdown(
                f"""
                <div class="feature-card">
                    <div style="font-size:28px;">{icon}</div>
                    <div style="font-weight:800; color:#0f2a5f; margin-top:8px;">{title}</div>
                    <div style="color:#64748b; font-size:12.5px; margin-top:6px;">{desc}</div>
                </div>
                """,
                unsafe_allow_html=True
            )

    st.markdown(
        '<div class="section-title">🛠️ What\'s included</div>',
        unsafe_allow_html=True
    )

    st.markdown("""
    <div class="info-box" style="animation: fadeInUp 0.6s ease .15s backwards;">
    <ul>
        <li>Manual column-mapping override for non-standard exports</li>
        <li>Data quality checks</li>
        <li>Smart match suggestions</li>
        <li>Vendor-wise ITC-at-risk summary</li>
        <li>Search + CSV export</li>
        <li>Professionally formatted Excel report</li>
        <li>GST Portal B2B + CDN + RCM analysis</li>
    </ul>
    </div>
    """, unsafe_allow_html=True)
